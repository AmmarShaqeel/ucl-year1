import openpyxl, pprint, math, output
from openpyxl import Workbook


wb = Workbook()
ws = wb.active

# defining constants
COIL_LENGTH = 14
COIL_HEIGHT = 13.5
COIL_DIAMETER = 27
RESISTIVTY = 1.7e-8
VOlTAGE = 1.5
C_RESISTANCE = 0.150
AAA_RESISTANCE = 0.250

# setting storage
storage = 0.0
biggest = [None]*5

# importing data from output.py
wireData = output.allData

ws.cell(row=1, column=1).value = "Gauge"
ws.cell(row=1, column=2).value = "AAA Parallel"
ws.cell(row=1, column=3).value = "AAA Series"
ws.cell(row=1, column=4).value = "C Parallel"
ws.cell(row=1, column=5).value = "C Series"

for i in wireData:
# calculating max layers + number of turns
	full = wireData[i]['spec']['full']
	copper = wireData[i]['spec']['copper']
	turns = COIL_LENGTH/full
	maxLayers =  math.floor(COIL_HEIGHT/full)
	maxTurns = maxLayers * turns
	
	wireData[i]['baseData'] = {'turns' : turns, 'maxLayers' : maxLayers, 'maxTurns' : maxTurns}
	
	wireData[i]['resistances'] = {}
	wireData[i]['currentsAAA_parallel'] = {}
	wireData[i]['currentsAAA_series'] = {}
	wireData[i]['currentsC_parallel'] = {}
	wireData[i]['currentsC_series'] = {}
	wireData[i]['maxLength'] = {}
	
	

	wireLength = turns * math.pi * ( (COIL_DIAMETER * 1e-3 * maxLayers) + ((maxLayers-1)*2)* full * 1e-3)  # arithmetic progression way of calculating lenght
	# wireLength = maxTurns * 2* math.pi * 0.0195
	resistance = RESISTIVTY * wireLength/(math.pi * (copper * 1e-3/2)** 2)
	
	
	currentsAAA_parallel = 1.5/(resistance + AAA_RESISTANCE/3)
	currentsAAA_series = 4.5/(resistance + AAA_RESISTANCE * 3)
	currentsC_parallel = 1.5/(resistance + C_RESISTANCE/3)
	currentsC_series = 4.5/(resistance + C_RESISTANCE * 3)
	
	ws.cell(row=i, column=1).value = i
	ws.cell(row=i, column=2).value = currentsAAA_parallel * turns * maxLayers
	ws.cell(row=i, column=3).value = currentsAAA_series * turns * maxLayers
	ws.cell(row=i, column=4).value = currentsC_parallel * turns * maxLayers
	ws.cell(row=i, column=5).value = currentsC_series * turns * maxLayers
	
	
	# keeping track of the biggest value of NI		
	if ((currentsAAA_parallel * turns * maxLayers) > storage):
		storage = currentsAAA_parallel * turns * maxLayers
		biggest[0] = currentsAAA_parallel * turns * maxLayers
		biggest[1] = i
		biggest[2] = turns
		biggest[3] = maxLayers
		biggest[4] = 'current AAA (parallel)'
		
	
	if ((currentsAAA_series * turns * maxLayers) > storage):
		storage = currentsAAA_series * turns * maxLayers
		biggest[0] = currentsAAA_series * turns * maxLayers
		biggest[1] = i
		biggest[2] = turns
		biggest[3] = maxLayers
		biggest[4] = 'current AAA (series)'
	
	if ((currentsC_parallel * turns * maxLayers) > storage):
		storage = currentsC_parallel * turns * maxLayers
		biggest[0] = currentsC_parallel * turns * maxLayers
		biggest[1] = i
		biggest[2] = turns
		biggest[3] = maxLayers
		biggest[4] = 'current C (parallel)'
	
	 
	if ((currentsC_series * turns * maxLayers) > storage):
		storage = currentsC_series * turns * maxLayers
		biggest[0] = currentsC_series * turns * maxLayers
		biggest[1] = i
		biggest[2] = turns
		biggest[3] = maxLayers
		biggest[4] = 'current C (series)'
	
	
	# storing data in dictionaries
	wireData[i]['resistances'] = resistance
	wireData[i]['currentsAAA_parallel'] = currentsAAA_parallel * turns * maxLayers
	wireData[i]['currentsAAA_series'] = currentsAAA_series * turns * maxLayers
	wireData[i]['currentsC_parallel'] = currentsC_parallel * turns * maxLayers
	wireData[i]['currentsC_series'] = currentsC_series * turns * maxLayers
	wireData[i]['maxLength'] = wireLength
		
		
		
		
print(biggest)		

	
# print(wireData[25]['currentsC_series'][1]*wireData[25]['baseData']['turns'])	
	
	
print('Writing results...')
resultFile = open('allVal.py', 'w')
resultFile.write('allData = ' + pprint.pformat(wireData))
resultFile.close()
print('Done.')

wb.save("sample.xlsx")		